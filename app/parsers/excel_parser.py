"""
Structure-aware Excel parser for benefit-related spreadsheets.

Goals:
- Detect table-like regions adaptively (no fixed template assumptions).
- Handle multi-row headers and irregular layouts.
- Produce semantic, human-readable sentences suitable for RAG.
- Degrade gracefully when structure is messy.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple


@dataclass
class ParsedTable:
    sheet: str
    section_title: Optional[str]
    table_id: int
    headers: List[str]
    rows: List[Dict[str, Any]]  # header -> value


def _is_non_empty(value: Any) -> bool:
    return value not in (None, "")


def _iter_non_empty_cells(ws) -> List[Tuple[int, int]]:
    """Return list of (row, col) for non-empty cells."""
    occupied: List[Tuple[int, int]] = []
    for row in ws.iter_rows():
        for cell in row:
            if _is_non_empty(cell.value):
                occupied.append((cell.row, cell.column))
    return occupied


def _infer_tables_in_sheet(ws) -> List[Tuple[int, int, int, int]]:
    """
    Detect candidate tables as rectangular regions:
    (start_row, end_row, min_col, max_col) per sheet.
    """
    occupied = _iter_non_empty_cells(ws)
    if not occupied:
        return []

    # Group contiguous rows with at least MIN_NONEMPTY_PER_ROW non-empty cells.
    MIN_NONEMPTY_PER_ROW = 2

    rows_nonempty: Dict[int, int] = {}
    for r, c in occupied:
        rows_nonempty[r] = rows_nonempty.get(r, 0) + 1

    tables_row_ranges: List[Tuple[int, int]] = []
    current_start: Optional[int] = None
    current_end: Optional[int] = None

    max_row = ws.max_row or 0
    for r in range(1, max_row + 1):
        count = rows_nonempty.get(r, 0)
        if count >= MIN_NONEMPTY_PER_ROW:
            if current_start is None:
                current_start = r
            current_end = r
        else:
            if current_start is not None:
                tables_row_ranges.append((current_start, current_end or current_start))
                current_start = None
                current_end = None
    if current_start is not None:
        tables_row_ranges.append((current_start, current_end or current_start))

    # For each row-range block, compute min/max column to form rectangle.
    tables: List[Tuple[int, int, int, int]] = []
    for start_row, end_row in tables_row_ranges:
        cols_in_block = [c for (r, c) in occupied if start_row <= r <= end_row]
        if not cols_in_block:
            continue
        min_col, max_col = min(cols_in_block), max(cols_in_block)
        tables.append((start_row, end_row, min_col, max_col))

    return tables


def _infer_section_title(ws, start_row: int, min_col: int, max_col: int) -> Optional[str]:
    """
    Look a few rows above start_row for mostly-text rows and treat them as titles.
    """
    title_rows: List[str] = []
    for r in range(max(start_row - 3, 1), start_row):
        texts: List[str] = []
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if not _is_non_empty(cell.value):
                continue
            v = str(cell.value).strip()
            if v:
                texts.append(v)
        if not texts:
            continue
        # Simple heuristic: treat as title if mostly non-numeric.
        digits = sum(ch.isdigit() for txt in texts for ch in txt)
        total = sum(1 for txt in texts for ch in txt)
        if total == 0 or digits < total:
            title_rows.append(" ".join(texts))
    if not title_rows:
        return None
    return " | ".join(title_rows)


def _detect_header_rows(ws, start_row: int, end_row: int, min_col: int, max_col: int) -> Tuple[List[int], int]:
    """
    Detect one or more header rows near top of the table.
    Returns (header_row_indices, data_start_row).
    """
    HEADER_MAX_ROWS = 3
    header_rows_idx: List[int] = []
    data_start_row = start_row

    for r in range(start_row, min(start_row + HEADER_MAX_ROWS, end_row + 1)):
        vals = [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
        non_empty = [v for v in vals if _is_non_empty(v)]
        if not non_empty:
            continue
        text_like = 0
        numeric_like = 0
        for v in non_empty:
            if isinstance(v, (int, float)):
                numeric_like += 1
            else:
                s = str(v)
                if any(ch.isdigit() for ch in s) and not any(ch.isalpha() for ch in s):
                    numeric_like += 1
                else:
                    text_like += 1
        if text_like >= numeric_like:
            header_rows_idx.append(r)
            data_start_row = r + 1
        else:
            break

    if not header_rows_idx:
        # Fallback: treat first row as header
        header_rows_idx = [start_row]
        data_start_row = start_row + 1

    return header_rows_idx, data_start_row


def _build_headers(
    ws,
    header_rows_idx: List[int],
    min_col: int,
    max_col: int,
) -> List[str]:
    """Merge multi-row headers into a single header label per column."""
    headers: List[str] = []
    for c in range(min_col, max_col + 1):
        parts: List[str] = []
        for r in header_rows_idx:
            cell = ws.cell(row=r, column=c)
            if not _is_non_empty(cell.value):
                continue
            s = str(cell.value).strip()
            if s:
                parts.append(s)
        if not parts:
            headers.append("")
        else:
            headers.append(" / ".join(parts))

    return headers


def _normalize_headers(headers: List[str]) -> Tuple[List[str], List[int]]:
    """
    Remove empty header names and return normalized headers and their original indices.
    """
    normalized: List[str] = []
    valid_indices: List[int] = []
    for i, h in enumerate(headers):
        h_clean = h.strip()
        if not h_clean:
            continue
        normalized.append(h_clean)
        valid_indices.append(i)
    return normalized, valid_indices


def _build_rows(
    ws,
    data_start_row: int,
    end_row: int,
    min_col: int,
    max_col: int,
    normalized_headers: List[str],
    valid_indices: List[int],
) -> List[Dict[str, Any]]:
    """Build row dicts using normalized headers."""
    rows: List[Dict[str, Any]] = []
    if not normalized_headers:
        return rows

    for r in range(data_start_row, end_row + 1):
        row_dict: Dict[str, Any] = {}
        for idx, c in enumerate(range(min_col, max_col + 1)):
            if idx not in valid_indices:
                continue
            header = normalized_headers[valid_indices.index(idx)]
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if not _is_non_empty(v):
                continue
            row_dict[header] = v
        if row_dict:
            rows.append(row_dict)
    return rows


def _parse_sheet_to_tables(ws) -> List[ParsedTable]:
    """Parse a single worksheet into ParsedTable objects."""
    tables_meta = _infer_tables_in_sheet(ws)
    tables: List[ParsedTable] = []

    table_id = 1
    for (start_row, end_row, min_col, max_col) in tables_meta:
        section_title = _infer_section_title(ws, start_row, min_col, max_col)
        header_rows_idx, data_start_row = _detect_header_rows(
            ws, start_row, end_row, min_col, max_col
        )
        headers = _build_headers(ws, header_rows_idx, min_col, max_col)
        normalized_headers, valid_indices = _normalize_headers(headers)
        rows = _build_rows(
            ws,
            data_start_row,
            end_row,
            min_col,
            max_col,
            normalized_headers,
            valid_indices,
        )
        if not rows:
            # Graceful degradation: skip empty/invalid tables
            continue

        tables.append(
            ParsedTable(
                sheet=ws.title,
                section_title=section_title,
                table_id=table_id,
                headers=normalized_headers,
                rows=rows,
            )
        )
        table_id += 1

    return tables


def _format_numeric(value: Any, header: str) -> str:
    """Format numeric values as currency/percent when appropriate."""
    if not isinstance(value, (int, float)):
        return str(value)

    h_lower = header.lower()
    if "percent" in h_lower or "%" in header:
        return f"{float(value):.2f}%"
    if any(word in h_lower for word in ["premium", "contribution", "rate", "cost", "fee"]):
        return f"${float(value):,.2f}"
    return str(value)


def _row_to_sentence(headers: List[str], row: Dict[str, Any]) -> str:
    """
    Turn a row dict into a human-readable sentence.
    Example:
      "Age 45: Employer premium is $144.40; Employee contribution is $33.32."
    """
    pieces: List[str] = []
    lead: Optional[str] = None

    for h in headers:
        if h not in row:
            continue
        val = row[h]
        val_str = _format_numeric(val, h)
        if lead is None:
            lead = f"{h} {val_str}"
        else:
            pieces.append(f"{h} is {val_str}")

    if lead is None:
        return ""

    if pieces:
        return f"{lead}: " + "; ".join(pieces) + "."
    return lead + "."


def _tables_to_semantic_lines(tables: List[ParsedTable]) -> List[str]:
    """
    Convert ParsedTable objects into semantic text lines suitable for RAG.
    Each line includes sheet/section/table context.
    """
    lines: List[str] = []
    for table in tables:
        headers = table.headers
        if not headers:
            continue
        for row in table.rows:
            sentence = _row_to_sentence(headers, row)
            if not sentence:
                continue
            prefix_parts = [f"Sheet {table.sheet}"]
            if table.section_title:
                prefix_parts.append(f"Section {table.section_title}")
            prefix_parts.append(f"Table {table.table_id}")
            prefix = " | ".join(prefix_parts)
            lines.append(f"{prefix}: {sentence}")
    return lines


def extract_text_from_excel_bytes(data: bytes) -> str:
    """
    Extract semantic text from an Excel workbook given as bytes.

    - Detects tables per sheet.
    - Normalizes (possibly multi-row) headers.
    - Converts each row into a human-readable sentence.
    - Prefixes each sentence with sheet/section/table context.
    - Degrades gracefully to a simple flat text fallback if parsing fails.
    """
    import io
    from openpyxl import load_workbook  # type: ignore[import]

    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        # Fallback: if workbook cannot be opened, return empty and let caller handle error.
        return ""

    all_tables: List[ParsedTable] = []
    for ws in wb.worksheets:
        try:
            tables = _parse_sheet_to_tables(ws)
            all_tables.extend(tables)
        except Exception:
            # Graceful degradation: skip problematic sheet
            continue

    if not all_tables:
        # Last-resort fallback: simple flat extraction like the original version.
        parts: List[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                row_vals = [cell.value for cell in row if _is_non_empty(cell.value)]
                if not row_vals:
                    continue
                parts.append(" | ".join(str(v).strip() for v in row_vals))
        return "\n".join(parts).strip()

    lines = _tables_to_semantic_lines(all_tables)
    return "\n".join(lines).strip()

